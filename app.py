import streamlit as st
import streamlit.components.v1 as components
import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Monthly Budget Review",
    layout="wide",
    initial_sidebar_state="collapsed",
)

NAVY = "#1e3a5f"

# ── Global styles ─────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
  html, body, [class*="css"] {{ font-family:'Inter',sans-serif; }}
  .stApp, .main {{ background:#f0f2f5 !important; }}
  .block-container {{ padding:0 !important; max-width:100% !important; }}
  #MainMenu, footer {{ visibility:hidden; }}
  [data-testid="stHeader"]  {{ display:none !important; }}
  [data-testid="stToolbar"] {{ display:none !important; }}

  .top-banner {{
    background:{NAVY}; padding:1rem 2.5rem;
    display:flex; align-items:center; justify-content:space-between;
  }}
  .top-banner h1 {{ color:#fff; font-size:1.1rem; font-weight:700; margin:0; }}
  .top-banner .sub {{ color:#93c5fd; font-size:0.85rem; font-weight:500; }}
  .top-banner .mode-badge {{
    background:rgba(255,255,255,0.15); color:#fff;
    padding:0.25rem 0.75rem; border-radius:999px;
    font-size:0.75rem; font-weight:600; letter-spacing:0.04em;
  }}
  .top-banner .sw-btn {{
    background:rgba(255,255,255,0.18); color:#fff;
    border:1px solid rgba(255,255,255,0.35); border-radius:8px;
    padding:0.28rem 0.9rem; font-size:0.78rem; font-weight:600;
    cursor:pointer; font-family:'Inter',sans-serif;
    transition:background 0.2s;
  }}
  .top-banner .sw-btn:hover {{ background:rgba(255,255,255,0.3); }}
  .inner {{ padding:2rem 2.5rem; }}
  .step-header {{ font-size:1.35rem; font-weight:700; color:#111827; margin-bottom:0.2rem; }}
  .step-sub    {{ font-size:0.88rem; color:#6b7280; margin-bottom:1.5rem; }}

  div[data-testid="stButton"] > button {{
    background:{NAVY}; color:white; border:none; border-radius:8px;
    padding:0.6rem 2rem; font-weight:600; font-size:0.9rem;
    transition:background 0.2s;
  }}
  div[data-testid="stButton"] > button:hover {{ background:#162d4a; }}

  label, div[data-testid="stWidgetLabel"] {{
    color:#111827 !important; font-size:0.875rem !important; font-weight:500 !important;
  }}
  input {{ color:#111827 !important; background:#f9fafb !important;
          border:1px solid #e5e7eb !important; border-radius:6px !important; }}
  .stSelectbox > div > div {{ color:#111827 !important; }}
  .stProgress > div > div > div > div {{ background:{NAVY} !important; }}

  /* Mode selector cards */
  .mode-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; margin-top:2rem; }}
  .mode-card {{
    background:#fff; border-radius:14px; padding:2rem 1.75rem;
    box-shadow:0 2px 8px rgba(0,0,0,0.08);
    border:2px solid transparent; cursor:pointer; transition:all 0.2s;
  }}
  .mode-card:hover {{ border-color:{NAVY}; box-shadow:0 4px 20px rgba(30,58,95,0.15); }}
  .mode-card h2 {{ color:{NAVY}; font-size:1.15rem; font-weight:700; margin-bottom:0.5rem; }}
  .mode-card p  {{ color:#6b7280; font-size:0.87rem; line-height:1.5; }}
  .mode-tag {{
    display:inline-block; margin-bottom:0.75rem;
    background:#eff6ff; color:{NAVY};
    padding:0.2rem 0.65rem; border-radius:999px;
    font-size:0.7rem; font-weight:700; letter-spacing:0.05em; text-transform:uppercase;
  }}

  /* White card for input sections */
  .stForm {{
    background-color: #ffffff !important;
    padding: 2rem !important;
    border-radius: 12px !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.07), 0 4px 16px rgba(0,0,0,0.05) !important;
    border: none !important;
  }}
  .col-headers {{
    display:grid; grid-template-columns:2fr 1fr 1fr;
    gap:0.75rem; margin-bottom:0.4rem;
  }}
  .col-headers span {{
    font-size:0.72rem; font-weight:600; color:#6b7280;
    text-transform:uppercase; letter-spacing:0.05em;
  }}
  .col-headers span.r {{ text-align:right; }}
</style>
""", unsafe_allow_html=True)

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]
YEARS  = [str(y) for y in range(2024, 2028)]

# ── Simple mode steps ─────────────────────────────────────────────────────────
SIMPLE_STEPS = [
    {"key":"income",   "title":"Income",
     "subtitle":"Enter all sources of monthly income."},
    {"key":"expenses", "title":"Monthly Expenses",
     "subtitle":"Enter your debt repayments, fixed costs, and variable spending."},
    {"key":"savings",  "title":"Savings & Investments",
     "subtitle":"Emergency fund contributions, retirement annuity, unit trusts, tax-free savings."},
]

# ── Complex mode sections ─────────────────────────────────────────────────────
COMPLEX_SECTIONS = [
    {
        "key": "income",
        "title": "Income",
        "subtitle": "Enter all sources of monthly income — projected and actual.",
        "fields": [
            ("salary",          "Salary / Primary Income"),
            ("additional",      "Additional Income"),
            ("rental",          "Rental Income"),
            ("business",        "Business Income"),
            ("other_income",    "Other Income"),
        ],
    },
    {
        "key": "debt",
        "title": "Debt Repayments",
        "subtitle": "All monthly debt obligations — minimum payments and contractual amounts.",
        "fields": [
            ("home_loan",       "Home Loan"),
            ("vehicle",         "Vehicle Finance"),
            ("credit_card",     "Credit Cards"),
            ("personal_loan",   "Personal Loans"),
            ("student_loan",    "Student Loans"),
            ("other_debt",      "Other Debt Repayments"),
        ],
    },
    {
        "key": "fixed",
        "title": "Fixed Expenses",
        "subtitle": "Recurring monthly bills that remain relatively constant.",
        "fields": [
            ("rent",            "Rent"),
            ("insurance",       "Insurance"),
            ("medical_aid",     "Medical Aid"),
            ("rates",           "Rates / Levies"),
            ("fibre",           "Fibre / Internet"),
            ("phone",           "Phone"),
            ("school_fees",     "School Fees"),
            ("subscriptions",   "Subscriptions"),
            ("security",        "Security"),
            ("domestic",        "Domestic Workers"),
            ("other_fixed",     "Other Fixed Expenses"),
        ],
    },
    {
        "key": "variable",
        "title": "Variable Expenses",
        "subtitle": "Spending that changes month to month based on lifestyle choices.",
        "fields": [
            ("groceries",       "Groceries"),
            ("fuel",            "Fuel / Transport"),
            ("electricity",     "Electricity / Gas"),
            ("clothing",        "Clothing"),
            ("entertainment",   "Entertainment"),
            ("restaurants",     "Restaurants / Takeaways"),
            ("alcohol",         "Alcohol / Cigarettes"),
            ("self_care",       "Self-Care / Personal Care"),
            ("travel",          "Holidays / Travel"),
            ("misc",            "Miscellaneous Discretionary"),
        ],
    },
    {
        "key": "savings",
        "title": "Savings & Investments",
        "subtitle": "Monthly contributions to wealth-building and emergency reserves. Savings and investments should be an essential item in everyone's budget.",
        "fields": [
            ("ra",              "Retirement Annuities"),
            ("unit_trusts",     "Unit Trusts"),
            ("tfsa",            "Tax-Free Savings Account"),
            ("offshore",        "Offshore Investments"),
            ("emergency",       "Emergency Fund Contributions"),
            ("other_savings",   "Other Investments / Savings"),
        ],
    },
]

CAT_COLORS_SIMPLE = {
    "Debt Repayments":       "#3b82f6",
    "Fixed Expenses":        "#8b5cf6",
    "Variable Expenses":     "#ec4899",
    "Savings & Investments": "#f59e0b",
}

CAT_COLORS_COMPLEX = {
    "Debt Repayments":   "#3b82f6",
    "Fixed Expenses":    "#8b5cf6",
    "Variable Expenses": "#ec4899",
    "Savings":           "#f59e0b",
}

COMPLEX_FIELD_COLORS = {
    "debt":     "#3b82f6",
    "fixed":    "#8b5cf6",
    "variable": "#ec4899",
    "savings":  "#f59e0b",
}

# ── Session state ─────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "page":         "mode_select",
        "mode":         None,
        "step":         0,
        "cx_step":      0,
        "month":        None,
        "year":         str(datetime.now().year),
        "data":         {},
        "cx_data":      {},
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

def fmt(v):  return f"R {int(round(v)):,}"
def pct(v):  return f"{v:.1f}%"

def zar_input(label, key, default=0.0):
    return st.number_input(label, min_value=0.0,
                           value=float(st.session_state.get(key, default)),
                           step=100.0, format="%.0f", key=key)

# ── Banner ────────────────────────────────────────────────────────────────────
def banner(sub="", show_switch=False, switch_key="", progress=None, progress_label=""):
    month_str = (f"{st.session_state.month} {st.session_state.year}"
                 if st.session_state.month else "")
    right_parts = []
    if month_str:
        right_parts.append(f'<span class="sub">{month_str}</span>')
    if st.session_state.mode:
        label = "Simple Mode" if st.session_state.mode == "simple" else "Complex Mode"
        right_parts.append(f'<span class="mode-badge">{label}</span>')
    right_html = "&nbsp;&nbsp;".join(right_parts) if right_parts else (
        f'<span class="sub">{sub}</span>' if sub else "")

    st.markdown(
        f'<div class="top-banner"><h1>Monthly Budget Review</h1>'
        f'<div style="display:flex;align-items:center;gap:0.75rem;">'
        f'{right_html}'
        f'</div></div>',
        unsafe_allow_html=True)

    if progress is not None:
        st.progress(progress)
        st.markdown(f"<p style='font-size:0.78rem;color:#374151;margin-top:-0.4rem;'>"
                    f"{progress_label}</p>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: Mode Selection
# ══════════════════════════════════════════════════════════════════════════════
def page_mode_select():
    banner("Select your budget mode")
    st.markdown('<div class="inner">', unsafe_allow_html=True)

    col = st.columns([1, 3, 1])[1]
    with col:
        st.markdown(
            f"<h2 style='color:{NAVY};font-weight:700;margin-bottom:0.25rem;'>"
            f"Select Budget Mode</h2>"
            f"<p style='color:#6b7280;font-size:0.9rem;margin-bottom:0;'>"
            f"Choose how detailed you want your monthly budget review to be.</p>",
            unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        c1, c2 = st.columns(2, gap="large")

        with c1:
            st.markdown(f"""
            <div class="mode-card">
              <h2>Simple Budget</h2>
              <p>A quick 6 step monthly check-in. Capture income, debt, fixed costs, variable spending, and savings.</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Select Simple Budget", use_container_width=True, key="btn_simple"):
                st.session_state.mode  = "simple"
                st.session_state.month = MONTHS[datetime.now().month - 1]
                st.session_state.year  = str(datetime.now().year)
                st.session_state.page  = "step"
                st.session_state.step  = 0
                st.rerun()

        with c2:
            st.markdown(f"""
            <div class="mode-card">
              <h2>Complex Budget</h2>
              <p>A comprehensive, line item categorisation. Income, debt, fixed costs, variable spending, and savings.</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Select Complex Budget", use_container_width=True, key="btn_complex"):
                st.session_state.mode    = "complex"
                st.session_state.month   = MONTHS[datetime.now().month - 1]
                st.session_state.year    = str(datetime.now().year)
                st.session_state.page    = "cx_step"
                st.session_state.cx_step = 0
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: Month selector (shared)
# ══════════════════════════════════════════════════════════════════════════════
def page_month():
    banner("Select a month to begin")
    st.markdown('<div class="inner">', unsafe_allow_html=True)

    _, col_sw, _ = st.columns([3, 1, 3])
    with col_sw:
        if st.button("Switch Mode", use_container_width=True, key="sw_month"):
            st.session_state.page = "mode_select"
            st.rerun()

    col = st.columns([1, 2, 1])[1]
    with col:
        st.markdown('<div class="step-header">Select Review Month</div>'
                    '<div class="step-sub">Choose the month and year you want to review.</div>',
                    unsafe_allow_html=True)
        month = st.selectbox("Month", MONTHS, index=datetime.now().month - 1, key="sel_month")
        year  = st.selectbox("Year",  YEARS,  index=YEARS.index(str(datetime.now().year)), key="sel_year")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Start Review", use_container_width=True, key="btn_start"):
            st.session_state.month = month
            st.session_state.year  = year
            if st.session_state.mode == "simple":
                st.session_state.page = "step"
                st.session_state.step = 0
            else:
                st.session_state.page    = "cx_step"
                st.session_state.cx_step = 0
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: Simple steps
# ══════════════════════════════════════════════════════════════════════════════

# Field definitions for each simple step
SIMPLE_INCOME_FIELDS = [
    ("income",      "Income (After Tax)"),
    ("income_add",  "Additional Incomes"),
]
SIMPLE_EXPENSE_FIELDS = [
    ("exp_debt",     "Debt Repayments"),
    ("exp_fixed",    "Fixed Expenses (Bills that stay the same e.g. rent)"),
    ("exp_variable", "Variable Expenses (Spending that changes month to month based on your lifestyle, e.g. groceries, restaurants)"),
]
SIMPLE_SAVINGS_FIELDS = [
    ("sav_savings",  "Savings (what you invest)"),
]

def _col_hint():
    h1, h2, h3 = st.columns([3, 1, 1])
    with h1: pass
    with h2:
        st.markdown(
            "<div style='font-size:0.7rem;font-weight:600;color:#6b7280;"
            "text-transform:uppercase;letter-spacing:0.05em;text-align:center;'>"
            "Projected<br><span style='font-weight:400;font-style:italic;"
            "text-transform:none;letter-spacing:0;'>What you think</span></div>",
            unsafe_allow_html=True)
    with h3:
        st.markdown(
            "<div style='font-size:0.7rem;font-weight:600;color:#6b7280;"
            "text-transform:uppercase;letter-spacing:0.05em;text-align:center;'>"
            "Actual<br><span style='font-weight:400;font-style:italic;"
            "text-transform:none;letter-spacing:0;'>What your bank statement says</span></div>",
            unsafe_allow_html=True)
    st.markdown("<div style='border-top:1px solid #e5e7eb;margin:0.3rem 0 0.5rem;'></div>",
                unsafe_allow_html=True)

def _field_row(label, fkey, data_store):
    p_prev = data_store.get(fkey, {}).get("proj", 0.0)
    a_prev = data_store.get(fkey, {}).get("actual", 0.0)
    c1, c2, c3 = st.columns([3, 1, 1])
    with c1:
        st.markdown(f"<p style='padding-top:0.45rem;color:#111827;font-size:0.88rem;"
                    f"margin:0;'>{label}</p>", unsafe_allow_html=True)
    with c2:
        p_val = st.number_input("", min_value=0.0, value=float(p_prev),
                                step=100.0, format="%.0f",
                                key=f"_s_{fkey}_p", label_visibility="collapsed")
    with c3:
        a_val = st.number_input("", min_value=0.0, value=float(a_prev),
                                step=100.0, format="%.0f",
                                key=f"_s_{fkey}_a", label_visibility="collapsed")
    return fkey, p_val, a_val

def page_simple_step():
    idx   = st.session_state.step
    total = len(SIMPLE_STEPS)
    s     = SIMPLE_STEPS[idx]

    banner(progress=(idx + 1) / total, progress_label=f"Step {idx+1} of {total}")
    st.markdown('<div class="inner">', unsafe_allow_html=True)

    data = st.session_state.data
    collected = {}

    _, mid, _ = st.columns([1, 4, 1])
    with mid:
        with st.form(key=f"simple_form_{idx}"):
            # ── Step 0: Income ────────────────────────────────────────────────
            if idx == 0:
                st.markdown(f'<div class="step-header">{s["title"]}</div>'
                            f'<div class="step-sub">{s["subtitle"]}</div>',
                            unsafe_allow_html=True)
                _col_hint()
                for fkey, flabel in SIMPLE_INCOME_FIELDS:
                    fk, pv, av = _field_row(flabel, fkey, data)
                    collected[fk] = {"proj": pv, "actual": av}
                sub_p = sum(v["proj"]   for v in collected.values())
                sub_a = sum(v["actual"] for v in collected.values())
                st.markdown(
                    f"<div style='border-top:2px solid {NAVY};margin-top:0.75rem;padding-top:0.6rem;"
                    f"display:grid;grid-template-columns:3fr 1fr 1fr;gap:0.75rem;'>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;'>Total</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_p)}</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_a)}</span></div>",
                    unsafe_allow_html=True)

            # ── Step 1: Expenses ──────────────────────────────────────────────
            elif idx == 1:
                st.markdown(f'<div class="step-header">{s["title"]}</div>'
                            f'<div class="step-sub">{s["subtitle"]}</div>',
                            unsafe_allow_html=True)
                _col_hint()
                for fkey, flabel in SIMPLE_EXPENSE_FIELDS:
                    fk, pv, av = _field_row(flabel, fkey, data)
                    collected[fk] = {"proj": pv, "actual": av}
                sub_p = sum(v["proj"]   for v in collected.values())
                sub_a = sum(v["actual"] for v in collected.values())
                st.markdown(
                    f"<div style='border-top:2px solid {NAVY};margin-top:0.75rem;padding-top:0.6rem;"
                    f"display:grid;grid-template-columns:3fr 1fr 1fr;gap:0.75rem;'>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;'>Total</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_p)}</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_a)}</span></div>",
                    unsafe_allow_html=True)

            # ── Step 2: Savings ───────────────────────────────────────────────
            elif idx == 2:
                st.markdown(f'<div class="step-header">{s["title"]}</div>'
                            f'<div class="step-sub">{s["subtitle"]}</div>',
                            unsafe_allow_html=True)
                _col_hint()
                for fkey, flabel in SIMPLE_SAVINGS_FIELDS:
                    fk, pv, av = _field_row(flabel, fkey, data)
                    collected[fk] = {"proj": pv, "actual": av}
                sub_p = sum(v["proj"]   for v in collected.values())
                sub_a = sum(v["actual"] for v in collected.values())
                st.markdown(
                    f"<div style='border-top:2px solid {NAVY};margin-top:0.75rem;padding-top:0.6rem;"
                    f"display:grid;grid-template-columns:3fr 1fr 1fr;gap:0.75rem;'>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;'>Total</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_p)}</span>"
                    f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;display:block;'>{fmt(sub_a)}</span></div>",
                    unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            nl, nr = st.columns(2)
            with nl:
                back_clicked = st.form_submit_button("Back" if idx > 0 else "Home", use_container_width=True)
            with nr:
                next_clicked = st.form_submit_button("Next" if idx < total - 1 else "View Dashboard", use_container_width=True, type="primary")

    if next_clicked:
        for fk, vals in collected.items():
            data[fk] = vals
        st.session_state.data = data
        if idx < total - 1:
            st.session_state.step += 1
        else:
            st.session_state.page = "dashboard"
        st.rerun()
    if back_clicked:
        for fk, vals in collected.items():
            data[fk] = vals
        st.session_state.data = data
        if idx > 0:
            st.session_state.step -= 1
        else:
            st.session_state.page = "mode_select"
        st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE: Complex section steps
# ══════════════════════════════════════════════════════════════════════════════
def page_complex_step():
    idx     = st.session_state.cx_step
    total   = len(COMPLEX_SECTIONS)
    section = COMPLEX_SECTIONS[idx]
    skey    = section["key"]
    has_toggle = skey in ("debt", "fixed", "variable")  # savings = always essential, income = no toggle

    banner(progress=(idx + 1) / total, progress_label=f"Section {idx+1} of {total}")
    st.markdown('<div class="inner">', unsafe_allow_html=True)

    # Centred narrow container
    _, mid, _ = st.columns([1, 4, 1])
    with mid:
        prev   = st.session_state.cx_data.get(skey, {})
        fields = section["fields"]

        with st.form(key=f"cx_form_{idx}"):
            st.markdown(f'<div class="step-header">{section["title"]}</div>'
                        f'<div class="step-sub">{section["subtitle"]}</div>',
                        unsafe_allow_html=True)

            # Column headers
            if has_toggle:
                hcols = st.columns([3, 1, 1, 1])
                labels = ["Line Item", "Essential?",
                          "Projected\nWhat you think", "Actual\nWhat your bank statement says"]
            else:
                hcols = st.columns([3, 1, 1])
                labels = ["Line Item",
                          "Projected\nWhat you think", "Actual\nWhat your bank statement says"]

            for hc, lbl in zip(hcols, labels):
                parts = lbl.split("\n")
                sub_part = f"<br><span style='font-weight:400;font-style:italic;text-transform:none;letter-spacing:0;font-size:0.65rem;'>{parts[1]}</span>" if len(parts) > 1 else ""
                with hc:
                    st.markdown(
                        f"<span style='font-size:0.72rem;font-weight:600;color:#6b7280;"
                        f"text-transform:uppercase;letter-spacing:0.05em;'>{parts[0]}{sub_part}</span>",
                        unsafe_allow_html=True)

            st.markdown("<div style='border-top:1px solid #e5e7eb;margin-bottom:0.5rem;'></div>",
                        unsafe_allow_html=True)

            field_vals = {}
            for fkey, flabel in fields:
                p_prev  = prev.get(fkey, {}).get("proj",      0.0)
                a_prev  = prev.get(fkey, {}).get("actual",    0.0)
                if skey == "savings":
                    ess_prev = True
                else:
                    ess_prev = prev.get(fkey, {}).get("essential", False)

                if has_toggle:
                    c1, c2, c3, c4 = st.columns([3, 1, 1, 1])
                else:
                    c1, c3, c4 = st.columns([3, 1, 1])

                with c1:
                    st.markdown(f"<p style='padding-top:0.45rem;color:#111827;font-size:0.88rem;"
                                f"margin:0;'>{flabel}</p>", unsafe_allow_html=True)
                if has_toggle:
                    with c2:
                        ess_val = st.toggle("", value=ess_prev,
                                            key=f"cx_{skey}_{fkey}_ess",
                                            label_visibility="collapsed")
                else:
                    ess_val = True

                with c3:
                    p_val = st.number_input("", min_value=0.0, value=float(p_prev),
                                            step=100.0, format="%.0f",
                                            key=f"cx_{skey}_{fkey}_p", label_visibility="collapsed")
                with c4:
                    a_val = st.number_input("", min_value=0.0, value=float(a_prev),
                                            step=100.0, format="%.0f",
                                            key=f"cx_{skey}_{fkey}_a", label_visibility="collapsed")
                field_vals[fkey] = {"proj": p_val, "actual": a_val, "essential": ess_val}

            # Subtotal
            sub_p = sum(v["proj"]   for v in field_vals.values())
            sub_a = sum(v["actual"] for v in field_vals.values())
            grid_cols = "3fr 1fr 1fr 1fr" if has_toggle else "3fr 1fr 1fr"
            st.markdown(
                f"<div style='border-top:2px solid {NAVY};margin-top:0.75rem;padding-top:0.6rem;"
                f"display:grid;grid-template-columns:{grid_cols};gap:0.75rem;'>"
                f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;'>Section Total</span>"
                + ("<span></span>" if has_toggle else "") +
                f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;"
                f"display:block;'>{fmt(sub_p)}</span>"
                f"<span style='font-weight:700;color:{NAVY};font-size:0.88rem;text-align:right;"
                f"display:block;'>{fmt(sub_a)}</span>"
                f"</div>",
                unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)
            nl, nr = st.columns(2)
            with nl:
                back_clicked = st.form_submit_button("Back" if idx > 0 else "Home", use_container_width=True)
            with nr:
                next_clicked = st.form_submit_button("Next" if idx < total - 1 else "View Dashboard", use_container_width=True, type="primary")

        if next_clicked:
            st.session_state.cx_data[skey] = field_vals
            if idx < total - 1: st.session_state.cx_step += 1
            else:                st.session_state.page = "cx_dashboard"
            st.rerun()
        if back_clicked:
            st.session_state.cx_data[skey] = field_vals
            if idx > 0: st.session_state.cx_step -= 1
            else:        st.session_state.page = "mode_select"
            st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# Simple Dashboard
# ══════════════════════════════════════════════════════════════════════════════
def compute_simple():
    d = st.session_state.data
    def g(k, t): return d.get(k, {}).get(t, 0)

    # Income
    tip = sum(g(k, "proj")   for k, _ in SIMPLE_INCOME_FIELDS)
    tia = sum(g(k, "actual") for k, _ in SIMPLE_INCOME_FIELDS)

    # Expenses (single line items)
    dp = g("exp_debt",     "proj");  da = g("exp_debt",     "actual")
    fp = g("exp_fixed",    "proj");  fa = g("exp_fixed",    "actual")
    vp = g("exp_variable", "proj");  va = g("exp_variable", "actual")

    # Savings
    sp = sum(g(k, "proj")   for k, _ in SIMPLE_SAVINGS_FIELDS)
    sa = sum(g(k, "actual") for k, _ in SIMPLE_SAVINGS_FIELDS)

    top_ = dp + fp + vp + sp
    toa  = da + fa + va + sa
    return dict(dp=dp,da=da,fp=fp,fa=fa,vp=vp,va=va,sp=sp,sa=sa,
                tip=tip,tia=tia,top_=top_,toa=toa,
                surp_p=tip-top_,surp_a=tia-toa)


def build_simple_dashboard(c, month_str):
    def vc(v): return "#059669" if v >= 0 else "#dc2626"
    def vs(v):
        sign = "+" if v >= 0 else "-"
        return f"{sign}R {int(round(abs(v))):,}"

    surplus_a   = c["surp_a"]
    total_inc_a = c["tia"]
    if total_inc_a == 0:
        pb,pf,pt = "#ffedd5","#9a3412","No Income Entered"
    elif surplus_a >= 0:
        pb,pf,pt = "#d1fae5","#065f46","Surplus"
    elif surplus_a / total_inc_a >= -0.05:
        pb,pf,pt = "#ffedd5","#9a3412","Tight — Small Deficit"
    else:
        pb,pf,pt = "#fee2e2","#991b1b","Deficit — Action Needed"

    exp_map = {"Debt Repayments":c["da"],"Fixed Expenses":c["fa"],
               "Variable Expenses":c["va"],"Savings & Investments":c["sa"]}
    largest = max(exp_map, key=exp_map.get)
    var_text = (f"{abs((c['va']-c['vp'])/c['vp']*100):.1f}% "
                f"{'over' if c['va']>c['vp'] else 'under'} target"
                if c["vp"] > 0 else "No target set")
    sav_text = (f"{c['sa']/total_inc_a*100:.1f}% of income"
                if total_inc_a > 0 else "—")
    req_red  = max(0, -surplus_a)
    req_txt  = fmt(req_red) if req_red > 0 else "None needed"
    req_col  = "#dc2626" if req_red > 0 else "#059669"
    net_col  = vc(surplus_a)

    pie_data = json.dumps({
        "labels": ["Debt Repayments","Fixed Expenses","Variable Expenses","Savings & Investments"],
        "values": [c["da"],c["fa"],c["va"],c["sa"]],
        "colors": ["#3b82f6","#8b5cf6","#ec4899","#f59e0b"],
        "text":   [
            "Home loan, vehicle, credit cards",
            "Rent, insurance, medical aid, subscriptions",
            "Groceries, fuel, dining, entertainment",
            "Retirement, unit trusts, tax-free savings"
        ],
    })
    bar_data = json.dumps({
        "cats":   ["Debt Repayments","Fixed Expenses","Variable Expenses","Savings & Investments"],
        "proj":   [c["dp"],c["fp"],c["vp"],c["sp"]],
        "actual": [c["da"],c["fa"],c["va"],c["sa"]],
    })

    def trow(label, p, a, bold=False, top_border=False, section=False, row_type="expense"):
        if not section and not bold and not top_border and p == 0 and a == 0:
            return ""
        v  = a - p
        if row_type == "savings":
            col = "#111827"
        elif row_type == "income":
            col = "#059669" if v >= 0 else "#dc2626"
        elif row_type == "net":
            col = "#059669" if v >= 0 else "#dc2626"
        else:  # expense: overspend (actual > projected) is bad = red
            col = "#dc2626" if v > 0 else "#059669"
        vstr = vs(v)
        w  = "font-weight:700;" if bold else ""
        tb = f"border-top:2px solid {NAVY};" if top_border else ""
        if section:
            return (f'<tr><td colspan="4" style="padding:0.65rem 0.75rem 0.2rem;'
                    f'font-size:0.78rem;font-weight:700;color:{NAVY};'
                    f'text-transform:uppercase;background:#fff;border-bottom:none;">'
                    f'{label}</td></tr>')
        return (f'<tr style="background:#fff;">'
                f'<td style="padding:0.45rem 0.75rem;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{label}</td>'
                f'<td style="padding:0.45rem 0.75rem;text-align:right;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{fmt(p)}</td>'
                f'<td style="padding:0.45rem 0.75rem;text-align:right;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{fmt(a)}</td>'
                f'<td style="padding:0.45rem 0.75rem;text-align:right;color:{col};font-weight:600;border-bottom:1px solid #f3f4f6;{tb}">{vstr}</td>'
                f'</tr>')

    def srow(label, val_html):
        return (f'<tr style="background:#fff;">'
                f'<td style="padding:0.5rem 0;color:#6b7280;border-bottom:1px solid #f3f4f6;font-size:0.84rem;">{label}</td>'
                f'<td style="padding:0.5rem 0;text-align:right;border-bottom:1px solid #f3f4f6;">{val_html}</td>'
                f'</tr>')

    table_rows = (
        trow("Income",0,0,section=True) +
        trow("Total Income",c["tip"],c["tia"],bold=True,row_type="income") +
        trow("Outflows",0,0,section=True) +
        trow("Debt Repayments",       c["dp"],c["da"],row_type="expense") +
        trow("Fixed Expenses",        c["fp"],c["fa"],row_type="expense") +
        trow("Variable Expenses",     c["vp"],c["va"],row_type="expense") +
        trow("Savings & Investments", c["sp"],c["sa"],row_type="savings") +
        trow("Total Outflows",c["top_"],c["toa"],bold=True,row_type="expense") +
        trow("Net Surplus / Deficit",c["surp_p"],c["surp_a"],bold=True,top_border=True,row_type="net")
    )

    summary_rows = (
        srow("What you should have left", f'<span style="font-weight:700;color:{net_col};">{fmt(surplus_a)}</span>') +
        srow("Variable Spending", f'<span style="font-weight:700;color:#111827;">{var_text}</span>') +
        srow("Savings Rate", f'<span style="font-weight:700;color:#111827;">{sav_text}</span>') +
        srow("Largest Expense Category", f'<span style="font-weight:700;color:#111827;">{largest}</span>') +
        srow("Required Reduction to Break Even", f'<span style="font-weight:700;color:{req_col};">{req_txt}</span>')
    )

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Inter',sans-serif;background:#f0f2f5;}}
  .banner{{background:{NAVY};padding:0.9rem 2rem;display:flex;align-items:center;justify-content:space-between;}}
  .banner h1{{color:#fff;font-size:1.05rem;font-weight:700;}}
  .banner .right{{display:flex;align-items:center;gap:0.75rem;}}
  .banner .sub{{color:#93c5fd;font-size:0.85rem;}}
  .banner .badge{{background:rgba(255,255,255,0.15);color:#fff;padding:0.2rem 0.65rem;border-radius:999px;font-size:0.72rem;font-weight:700;letter-spacing:0.04em;}}
  .controls{{padding:0.75rem 2rem 0;display:flex;justify-content:flex-end;gap:0.75rem;}}
  .btn{{background:{NAVY};color:#fff;border:none;border-radius:8px;padding:0.4rem 1.1rem;font-size:0.82rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif;}}
  .btn:hover{{background:#162d4a;}}
  .grid{{display:grid;grid-template-columns:1fr 1fr;grid-auto-rows:1fr;gap:1.25rem;padding:1rem 2rem 2rem;}}
  .card{{background:#ffffff;border-radius:12px;padding:1.5rem;box-shadow:0 1px 4px rgba(0,0,0,0.08),0 4px 16px rgba(0,0,0,0.06);display:flex;flex-direction:column;}}
  .card-title{{font-size:0.7rem;font-weight:600;letter-spacing:0.08em;text-transform:uppercase;color:#6b7280;margin-bottom:1rem;flex-shrink:0;}}
  .budget-table{{width:100%;border-collapse:collapse;font-size:0.84rem;}}
  .budget-table thead th{{background:{NAVY};color:#fff;padding:0.5rem 0.75rem;text-align:left;font-size:0.72rem;font-weight:600;letter-spacing:0.06em;text-transform:uppercase;}}
  .budget-table thead th.r{{text-align:right;}}
  .sum-table{{width:100%;border-collapse:collapse;font-size:0.84rem;}}
  .sum-table td{{padding:0.5rem 0;border-bottom:1px solid #f3f4f6;vertical-align:middle;}}
  .sum-table tr:last-child td{{border-bottom:none;}}
  .pill{{display:inline-block;padding:0.28rem 0.8rem;border-radius:999px;font-size:0.76rem;font-weight:600;margin-bottom:0.85rem;background:{pb};color:{pf};}}
  .chart-wrap{{flex:1;min-height:180px;}}
  @media print{{
    body{{background:#fff;}}
    .banner{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
    .card{{box-shadow:none;border:1px solid #e5e7eb;page-break-inside:avoid;}}
    .grid{{gap:0.75rem;}}
  }}
</style></head><body>
<div class="banner">
  <h1>Monthly Budget Review</h1>
  <div class="right"><span class="sub">{month_str}</span><span class="badge">Simple Mode</span></div>
</div>
<div class="grid">
  <div class="card">
    <div class="card-title">Projected vs Actual</div>
    <table class="budget-table">
      <thead><tr><th>Category</th><th class="r">Projected</th><th class="r">Actual</th><th class="r">Variance</th></tr></thead>
      <tbody>{table_rows}</tbody>
    </table>
  </div>
  <div class="card">
    <div class="card-title">Actual Expense Breakdown</div>
    <div class="chart-wrap" id="pie"></div>
  </div>
  <div class="card">
    <div class="card-title">Budget Summary</div>
    <div class="pill">{pt}</div>
    <table class="sum-table"><tbody>{summary_rows}</tbody></table>
  </div>
  <div class="card">
    <div class="card-title">Projected vs Actual by Category</div>
    <div class="chart-wrap" id="bar"></div>
  </div>
</div>
<script>
var pd={pie_data}, bd={bar_data};
Plotly.newPlot('pie',[{{
  type:'pie',
  labels:pd.labels,
  values:pd.values,
  hole:0.52,
  marker:{{colors:pd.colors,line:{{color:'#fff',width:2}}}},
  textinfo:'label+percent',
  texttemplate:'<b>%{{label}}</b><br>%{{percent:.1%}}',
  textposition:'outside',
  textfont:{{size:13,color:'#1f2937',family:'Inter'}},
  customdata:pd.text,
  hovertemplate:'<b>%{{label}}</b><br>%{{customdata}}<br>R %{{value:,}}<br>%{{percent:.1%}}<extra></extra>',
  automargin:true,
  pull:[0.03,0.03,0.03,0.03],
}}],
  {{paper_bgcolor:'#fff',plot_bgcolor:'#fff',
    font:{{family:'Inter',color:'#1f2937',size:13}},
    margin:{{l:60,r:60,t:40,b:40}},
    showlegend:false,
    autosize:true}},
  {{responsive:true,displayModeBar:false}});
Plotly.newPlot('bar',[
  {{name:'Projected',type:'bar',x:bd.cats,y:bd.proj,marker:{{color:'#93c5fd'}}}},
  {{name:'Actual',type:'bar',x:bd.cats,y:bd.actual,marker:{{color:'{NAVY}'}}}}],
  {{paper_bgcolor:'#fff',plot_bgcolor:'#fff',font:{{family:'Inter',color:'#374151',size:11}},
    margin:{{l:70,r:10,t:30,b:60}},barmode:'group',bargap:0.25,bargroupgap:0.08,
    xaxis:{{showgrid:false,linecolor:'#e5e7eb',tickfont:{{size:11}}}},
    yaxis:{{showgrid:true,gridcolor:'#f3f4f6',tickprefix:'R ',tickformat:',d'}},
    legend:{{orientation:'h',x:0,y:1.08,font:{{size:11}}}},autosize:true}},
  {{responsive:true,displayModeBar:false}});
</script></body></html>"""


def page_simple_dashboard():
    c = compute_simple()
    month_str = f"{st.session_state.month} {st.session_state.year}"

    col_sw, col_ed, col_xl, col_pdf, col_so = st.columns([2, 1, 1, 1, 1])
    with col_sw:
        if st.button("Switch Mode", use_container_width=True, key="sw_sd"):
            st.session_state.page = "mode_select"
            st.rerun()
    with col_ed:
        if st.button("Edit", use_container_width=True, key="ed_sd"):
            st.session_state.page = "step"
            st.session_state.step = 0
            st.rerun()
    with col_xl:
        buf = build_simple_excel(c, month_str)
        st.download_button("Download Excel", data=buf,
                           file_name=f"budget_{month_str.replace(' ','_')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="xl_sd")
    with col_pdf:
        pdf_buf = build_simple_pdf(c, month_str)
        st.download_button("Download PDF", data=pdf_buf,
                           file_name=f"budget_{month_str.replace(' ', '_')}.pdf",
                           mime="application/pdf",
                           use_container_width=True, key="pdf_sd")
    with col_so:
        if st.button("Start Over", use_container_width=True, key="so_sd"):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()

    html = build_simple_dashboard(c, month_str)
    components.html(html, height=1050, scrolling=False)

# ══════════════════════════════════════════════════════════════════════════════
# Complex Dashboard
# ══════════════════════════════════════════════════════════════════════════════
def compute_complex():
    d = st.session_state.cx_data

    def field(skey, fkey, t):
        return d.get(skey, {}).get(fkey, {}).get(t, 0)

    def is_essential(skey, fkey):
        # savings always essential; income irrelevant; others use toggle
        if skey in ("savings", "income"):
            return True
        return d.get(skey, {}).get(fkey, {}).get("essential", True)

    # Income
    inc_fields = ["salary","additional","rental","business","other_income"]
    inc_p = {f: field("income", f, "proj")   for f in inc_fields}
    inc_a = {f: field("income", f, "actual") for f in inc_fields}
    tip = sum(inc_p.values())
    tia = sum(inc_a.values())

    # Debt
    debt_fields = ["home_loan","vehicle","credit_card","personal_loan","student_loan","other_debt"]
    debt_p = {f: field("debt", f, "proj")   for f in debt_fields}
    debt_a = {f: field("debt", f, "actual") for f in debt_fields}
    tdp = sum(debt_p.values()); tda = sum(debt_a.values())

    # Fixed
    fixed_fields = ["rent","insurance","medical_aid","rates","fibre","phone",
                    "school_fees","subscriptions","security","domestic","other_fixed"]
    fixed_p = {f: field("fixed", f, "proj")   for f in fixed_fields}
    fixed_a = {f: field("fixed", f, "actual") for f in fixed_fields}
    tfp = sum(fixed_p.values()); tfa = sum(fixed_a.values())

    # Variable
    var_fields = ["groceries","fuel","electricity","clothing","entertainment",
                  "restaurants","alcohol","self_care","travel","misc"]
    var_p = {f: field("variable", f, "proj")   for f in var_fields}
    var_a = {f: field("variable", f, "actual") for f in var_fields}
    tvp = sum(var_p.values()); tva = sum(var_a.values())

    # Savings (always essential)
    sav_fields = ["ra","unit_trusts","tfsa","offshore","emergency","other_savings"]
    sav_p = {f: field("savings", f, "proj")   for f in sav_fields}
    sav_a = {f: field("savings", f, "actual") for f in sav_fields}
    tsp = sum(sav_p.values()); tsa = sum(sav_a.values())

    total_out_p = tdp + tfp + tvp + tsp
    total_out_a = tda + tfa + tva + tsa

    # Essential / Non-Essential actual totals (savings always essential)
    ess_a = tsa  # start with savings
    non_a = 0.0
    for skey, fkeys in [("debt", debt_fields), ("fixed", fixed_fields), ("variable", var_fields)]:
        for fk in fkeys:
            val = field(skey, fk, "actual")
            if is_essential(skey, fk):
                ess_a += val
            else:
                non_a += val

    ess_p = tsp  # savings projected
    non_p = 0.0
    for skey, fkeys in [("debt", debt_fields), ("fixed", fixed_fields), ("variable", var_fields)]:
        for fk in fkeys:
            val = field(skey, fk, "proj")
            if is_essential(skey, fk):
                ess_p += val
            else:
                non_p += val

    return dict(
        tip=tip, tia=tia,
        inc_p=inc_p, inc_a=inc_a,
        debt_p=debt_p, debt_a=debt_a, tdp=tdp, tda=tda,
        fixed_p=fixed_p, fixed_a=fixed_a, tfp=tfp, tfa=tfa,
        var_p=var_p, var_a=var_a, tvp=tvp, tva=tva,
        sav_p=sav_p, sav_a=sav_a, tsp=tsp, tsa=tsa,
        total_out_p=total_out_p, total_out_a=total_out_a,
        surp_p=tip - total_out_p,
        surp_a=tia - total_out_a,
        ess_p=ess_p, ess_a=ess_a,
        non_p=non_p, non_a=non_a,
    )


def build_complex_dashboard(c, month_str):
    NAVY2 = NAVY

    def vc(v):  return "#059669" if v >= 0 else "#dc2626"
    def vcn(v): return "#dc2626" if v > 0 else "#059669"  # for overspend (positive=bad)
    def vs(v, pct_base=None):
        sign = "+" if v >= 0 else "-"
        base = f"{sign}R {int(round(abs(v))):,}"
        if pct_base and pct_base > 0:
            p = abs(v / pct_base * 100)
            base += f" ({p:.1f}%)"
        return base

    # ── Ratios ────────────────────────────────────────────────────────────────
    tia  = c["tia"]
    surp = c["surp_a"]

    sav_rate   = (c["tsa"] / tia * 100)      if tia > 0 else 0
    dti        = (c["tda"] / tia * 100)      if tia > 0 else 0
    fixed_rat  = ((c["tda"]+c["tfa"]) / tia * 100) if tia > 0 else 0
    total_exp  = c["total_out_a"]
    ess_pct    = (c["ess_a"] / total_exp * 100) if total_exp > 0 else 0
    non_pct    = (c["non_a"] / total_exp * 100) if total_exp > 0 else 0
    req_red    = max(0, -surp)

    # Health score
    if tia == 0:
        health_col, health_txt = "#f59e0b", "Insufficient Data"
    elif surp >= 0 and dti < 30 and sav_rate >= 20:
        health_col, health_txt = "#059669", "Financially Healthy"
    elif surp >= 0 and (dti < 40 or sav_rate >= 10):
        health_col, health_txt = "#f59e0b", "Moderate Concern"
    else:
        health_col, health_txt = "#dc2626", "High Concern"

    # Net pill
    if tia == 0:
        pb,pf,pt = "#ffedd5","#9a3412","No Income Entered"
    elif surp >= 0:
        pb,pf,pt = "#d1fae5","#065f46","Surplus"
    elif surp / tia >= -0.05:
        pb,pf,pt = "#ffedd5","#9a3412","Tight — Small Deficit"
    else:
        pb,pf,pt = "#fee2e2","#991b1b","Deficit — Action Needed"

    # ── Table rows builder ────────────────────────────────────────────────────
    def sec_hdr(label):
        return (f'<tr><td colspan="5" style="padding:0.7rem 0.75rem 0.2rem;'
                f'font-size:0.78rem;font-weight:700;color:{NAVY2};'
                f'text-transform:uppercase;background:#fff;border-bottom:none;">'
                f'{label}</td></tr>')

    def drow(label, p, a, is_total=False, is_net=False, row_type="expense"):
        if not is_total and not is_net and p == 0 and a == 0:
            return ""
        v    = a - p
        if row_type == "savings":
            col = "#111827"
            pct_col = "#111827"
        elif row_type == "income":
            col = "#059669" if v >= 0 else "#dc2626"
            pct_col = col
        elif row_type == "net":
            col = "#059669" if v >= 0 else "#dc2626"
            pct_col = col
        else:  # expense: actual > projected = overspend = red
            col = "#dc2626" if v > 0 else "#059669"
            pct_col = col
        vstr = vs(v, p)
        w    = "font-weight:700;" if (is_total or is_net) else ""
        tb   = f"border-top:2px solid {NAVY2};" if is_net else ""
        pct_str = f"{abs(v/p*100):.1f}%" if p > 0 else "—"
        return (f'<tr style="background:#fff;">'
                f'<td style="padding:0.4rem 0.6rem;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{label}</td>'
                f'<td style="padding:0.4rem 0.6rem;text-align:right;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{fmt(p)}</td>'
                f'<td style="padding:0.4rem 0.6rem;text-align:right;color:#111827;border-bottom:1px solid #f3f4f6;{w}{tb}">{fmt(a)}</td>'
                f'<td style="padding:0.4rem 0.6rem;text-align:right;color:{col};font-weight:600;border-bottom:1px solid #f3f4f6;{tb}">{vs(v)}</td>'
                f'<td style="padding:0.4rem 0.6rem;text-align:right;color:{pct_col};font-weight:600;border-bottom:1px solid #f3f4f6;{tb}">{pct_str}</td>'
                f'</tr>')

    # Debt rows
    debt_label_map = {"home_loan":"Home Loan","vehicle":"Vehicle Finance",
                      "credit_card":"Credit Cards","personal_loan":"Personal Loans",
                      "student_loan":"Student Loans","other_debt":"Other Debt"}
    fixed_label_map = {"rent":"Rent","insurance":"Insurance","medical_aid":"Medical Aid",
                       "rates":"Rates / Levies","fibre":"Fibre / Internet","phone":"Phone",
                       "school_fees":"School Fees","subscriptions":"Subscriptions",
                       "security":"Security","domestic":"Domestic Workers","other_fixed":"Other Fixed"}
    var_label_map = {"groceries":"Groceries","fuel":"Fuel / Transport",
                     "electricity":"Electricity / Gas","clothing":"Clothing",
                     "entertainment":"Entertainment","restaurants":"Restaurants / Takeaways",
                     "alcohol":"Alcohol / Cigarettes","self_care":"Self-Care",
                     "travel":"Holidays / Travel","misc":"Miscellaneous"}
    sav_label_map = {"ra":"Retirement Annuities","unit_trusts":"Unit Trusts",
                     "tfsa":"Tax-Free Savings","offshore":"Offshore Investments",
                     "emergency":"Emergency Fund","other_savings":"Other Savings"}

    inc_label_map = {"salary":"Salary / Primary Income","additional":"Additional Income",
                     "rental":"Rental Income","business":"Business Income","other_income":"Other Income"}

    def section_rows(label_map, p_dict, a_dict, total_p, total_a, section_label, row_type="expense"):
        html = sec_hdr(section_label)
        for fkey, flabel in label_map.items():
            html += drow(flabel, p_dict.get(fkey,0), a_dict.get(fkey,0), row_type=row_type)
        html += drow(f"Total {section_label}", total_p, total_a, is_total=True, row_type=row_type)
        return html

    inc_rows  = sec_hdr("Income")
    for fkey, flabel in inc_label_map.items():
        inc_rows += drow(flabel, c["inc_p"].get(fkey,0), c["inc_a"].get(fkey,0), row_type="income")
    inc_rows += drow("Total Income", c["tip"], c["tia"], is_total=True, row_type="income")

    table_rows = (
        inc_rows +
        section_rows(debt_label_map,  c["debt_p"],  c["debt_a"],  c["tdp"], c["tda"],  "Debt Repayments",      row_type="expense") +
        section_rows(fixed_label_map, c["fixed_p"], c["fixed_a"], c["tfp"], c["tfa"],  "Fixed Expenses",       row_type="expense") +
        section_rows(var_label_map,   c["var_p"],   c["var_a"],   c["tvp"], c["tva"],  "Variable Expenses",    row_type="expense") +
        section_rows(sav_label_map,   c["sav_p"],   c["sav_a"],   c["tsp"], c["tsa"],  "Savings & Investments",row_type="savings") +
        drow("Net Surplus / Deficit", c["surp_p"], c["surp_a"], is_net=True, row_type="net")
    )

    # ── Pie data (inner=sections, outer=subcategories) ────────────────────────
    inner_labels = ["Debt","Fixed","Variable","Savings"]
    inner_values = [c["tda"], c["tfa"], c["tva"], c["tsa"]]
    inner_colors = ["#3b82f6","#8b5cf6","#ec4899","#f59e0b"]

    outer_labels, outer_values, outer_colors = [], [], []
    for fk, fl in debt_label_map.items():
        outer_labels.append(fl); outer_values.append(c["debt_a"].get(fk,0))
        outer_colors.append(["#93c5fd","#60a5fa","#3b82f6","#2563eb","#1d4ed8","#1e40af"][
            list(debt_label_map.keys()).index(fk) % 6])
    for fk, fl in fixed_label_map.items():
        outer_labels.append(fl); outer_values.append(c["fixed_a"].get(fk,0))
        outer_colors.append(["#c4b5fd","#a78bfa","#8b5cf6","#7c3aed","#6d28d9",
                              "#5b21b6","#4c1d95","#6d28d9","#7c3aed","#8b5cf6","#a78bfa"][
            list(fixed_label_map.keys()).index(fk) % 11])
    for fk, fl in var_label_map.items():
        outer_labels.append(fl); outer_values.append(c["var_a"].get(fk,0))
        outer_colors.append(["#fbcfe8","#f9a8d4","#f472b6","#ec4899","#db2777",
                              "#be185d","#9d174d","#831843","#be185d","#db2777"][
            list(var_label_map.keys()).index(fk) % 10])
    for fk, fl in sav_label_map.items():
        outer_labels.append(fl); outer_values.append(c["sav_a"].get(fk,0))
        outer_colors.append(["#fde68a","#fcd34d","#fbbf24","#f59e0b","#d97706","#b45309"][
            list(sav_label_map.keys()).index(fk) % 6])

    pie_inner = json.dumps({"labels":inner_labels,"values":inner_values,"colors":inner_colors})
    pie_outer = json.dumps({"labels":outer_labels,"values":outer_values,"colors":outer_colors})

    # ── Bar chart data ────────────────────────────────────────────────────────
    # ── Advisor insights ──────────────────────────────────────────────────────
    insights = []
    warnings_html = ""
    positives_html = ""

    if dti > 40:
        insights.append(("warning", f"Debt levels are very high at {dti:.1f}% of income and may significantly reduce financial flexibility."))
    elif dti > 30:
        insights.append(("caution", f"Debt levels are elevated at {dti:.1f}% of income and worth monitoring closely."))

    if sav_rate < 10 and tia > 0:
        insights.append(("warning", f"Savings are currently low at {sav_rate:.1f}% of income. Increasing this over time would improve long-term stability."))
    elif sav_rate >= 20:
        insights.append(("positive", f"Strong savings habits — {sav_rate:.1f}% of income is being saved consistently."))

    if surp < 0:
        insights.append(("warning", f"Monthly spending exceeds income by {fmt(abs(surp))}. This may not be sustainable long term."))
    elif surp > 0:
        insights.append(("positive", f"Monthly cashflow is positive with a surplus of {fmt(surp)}."))

    if fixed_rat > 60 and tia > 0:
        insights.append(("warning", "A large portion of income is tied up in fixed costs, which can reduce flexibility."))

    if non_pct > 30 and tia > 0:
        insights.append(("caution", f"Non-essential spending is relatively high at {non_pct:.1f}% of total expenses."))

    if c["var_a"].get("restaurants", 0) + c["var_a"].get("alcohol", 0) > tia * 0.1 and tia > 0:
        insights.append(("caution", "Spending on dining and entertainment is above average and may be an area to review."))

    if not insights:
        insights.append(("positive", "Overall spending and savings levels appear reasonably balanced."))

    def insight_row(itype, text):
        if itype == "warning":
            bg, dot = "#fee2e2", "#dc2626"
        elif itype == "caution":
            bg, dot = "#ffedd5", "#f59e0b"
        else:
            bg, dot = "#d1fae5", "#059669"
        return (f'<div style="display:flex;gap:0.6rem;align-items:flex-start;'
                f'background:{bg};border-radius:8px;padding:0.6rem 0.75rem;margin-bottom:0.5rem;">'
                f'<div style="width:8px;height:8px;border-radius:50%;background:{dot};'
                f'flex-shrink:0;margin-top:0.25rem;"></div>'
                f'<span style="font-size:0.82rem;color:#111827;line-height:1.45;">{text}</span></div>')

    insights_html = "".join(insight_row(t, tx) for t, tx in insights)

    def ratio_row(label, value_str, sub=""):
        sub_html = f'<br><span style="font-size:0.72rem;color:#9ca3af;">{sub}</span>' if sub else ""
        return (f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:0.5rem 0;border-bottom:1px solid #f3f4f6;">'
                f'<span style="color:#6b7280;font-size:0.83rem;">{label}{sub_html}</span>'
                f'<span style="font-weight:700;color:#111827;font-size:0.84rem;">{value_str}</span></div>')

    ratios_html = (
        ratio_row("Savings Rate", pct(sav_rate), "Target: >20%") +
        ratio_row("Debt-to-Income Ratio", pct(dti), "Safe zone: <30%") +
        ratio_row("Fixed Obligation Ratio", pct(fixed_rat), "Debt + fixed costs vs income") +
        ratio_row("Essential Spending", pct(ess_pct), "% of total expenses") +
        ratio_row("Non-Essential Spending", pct(non_pct), "% of total expenses") +
        ratio_row("Income Surplus", fmt(surp)) +
        ratio_row("Required Reduction to Break Even",
                  fmt(req_red) if req_red > 0 else "None needed")
    )

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Inter',sans-serif;background:#f0f2f5;}}
  .banner{{background:{NAVY2};padding:0.9rem 2rem;display:flex;align-items:center;justify-content:space-between;}}
  .banner h1{{color:#fff;font-size:1.05rem;font-weight:700;}}
  .banner .right{{display:flex;align-items:center;gap:0.75rem;}}
  .banner .sub{{color:#93c5fd;font-size:0.85rem;}}
  .banner .badge{{background:rgba(255,255,255,0.15);color:#fff;padding:0.2rem 0.65rem;border-radius:999px;font-size:0.72rem;font-weight:700;}}
  .controls{{padding:0.75rem 2rem 0;display:flex;justify-content:flex-end;gap:0.75rem;}}
  .btn{{background:{NAVY2};color:#fff;border:none;border-radius:8px;padding:0.4rem 1.1rem;font-size:0.82rem;font-weight:600;cursor:pointer;font-family:'Inter',sans-serif;}}
  .btn:hover{{background:#162d4a;}}
  .grid{{display:grid;grid-template-columns:1fr 1fr;gap:1.25rem;padding:1rem 2rem 2rem;}}
  .card{{background:#ffffff;border-radius:12px;padding:1.5rem;box-shadow:0 1px 4px rgba(0,0,0,0.08),0 4px 16px rgba(0,0,0,0.06);display:flex;flex-direction:column;}}
  .card-title{{font-size:0.7rem;font-weight:600;letter-spacing:0.08em;text-transform:uppercase;color:#6b7280;margin-bottom:1rem;flex-shrink:0;}}
  .budget-table{{width:100%;border-collapse:collapse;font-size:0.8rem;}}
  .budget-table thead th{{background:{NAVY2};color:#fff;padding:0.45rem 0.6rem;text-align:left;font-size:0.68rem;font-weight:600;letter-spacing:0.06em;text-transform:uppercase;}}
  .budget-table thead th.r{{text-align:right;}}
  .pill{{display:inline-block;padding:0.28rem 0.8rem;border-radius:999px;font-size:0.76rem;font-weight:600;margin-bottom:0.85rem;background:{pb};color:{pf};}}
  .health-badge{{display:inline-block;padding:0.3rem 0.9rem;border-radius:999px;font-size:0.78rem;font-weight:700;color:#fff;background:{health_col};margin-bottom:0.85rem;}}
  .chart-wrap{{flex:1;min-height:200px;}}
  .scroll-table{{flex:1;overflow-y:auto;max-height:520px;}}
  @media print{{
    body{{background:#fff;}}
    .banner{{-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
    .card{{box-shadow:none;border:1px solid #e5e7eb;page-break-inside:avoid;}}
    .grid{{gap:0.75rem;}}
    .scroll-table{{max-height:none;overflow:visible;}}
  }}
</style></head><body>
<div class="banner">
  <h1>Monthly Budget Review</h1>
  <div class="right"><span class="sub">{month_str}</span><span class="badge">Complex Mode</span></div>
</div>
<div class="grid">

  <!-- TOP LEFT: Detailed variance table -->
  <div class="card">
    <div class="card-title">Detailed Budget Variance</div>
    <div class="scroll-table">
    <table class="budget-table">
      <thead><tr>
        <th>Category</th>
        <th class="r">Projected</th>
        <th class="r">Actual</th>
        <th class="r">Variance</th>
        <th class="r">% Diff</th>
      </tr></thead>
      <tbody>{table_rows}</tbody>
    </table>
    </div>
  </div>

  <!-- TOP RIGHT: Donut chart (inner+outer) -->
  <div class="card">
    <div class="card-title">Expense Composition & Breakdown</div>
    <div class="chart-wrap" id="pie"></div>
  </div>

  <!-- BOTTOM LEFT: Advisor insights -->
  <div class="card">
    <div class="card-title">Advisor Insights & Financial Health</div>
    <div class="health-badge">{health_txt}</div>
    <div class="pill">Income Surplus: {fmt(surp)}</div>
    {insights_html}
  </div>

  <!-- BOTTOM RIGHT: Ratios -->
  <div class="card">
    <div class="card-title">Financial Ratios & Analytics</div>
    {ratios_html}
  </div>

</div>
<script>
var pi = {pie_inner};
var po = {pie_outer};

// Outer ring = major sections (labels outside), Inner ring = subcategories (hover only)
// sort:false on both traces keeps arc order matching data order so rings align
Plotly.newPlot('pie',[
  {{
    type:'pie',
    labels:po.labels,
    values:po.values,
    hole:0.3,
    sort:false,
    direction:'clockwise',
    marker:{{colors:po.colors,line:{{color:'#fff',width:1}}}},
    textinfo:'none',
    hovertemplate:'<b>%{{label}}</b><br>R %{{value:,d}}<extra></extra>',
    name:'Subcategories',
    domain:{{x:[0,1],y:[0,1]}}
  }},
  {{
    type:'pie',
    labels:pi.labels,
    values:pi.values,
    hole:0.58,
    sort:false,
    direction:'clockwise',
    marker:{{colors:pi.colors,line:{{color:'#fff',width:3}}}},
    textinfo:'label+percent',
    texttemplate:'<b>%{{label}}</b><br>%{{percent:.1%}}',
    textposition:'outside',
    textfont:{{size:13,color:'#1f2937',family:'Inter'}},
    hovertemplate:'<b>%{{label}}</b><br>R %{{value:,d}}<br>%{{percent:.1%}}<extra></extra>',
    automargin:true,
    name:'Sections',
    domain:{{x:[0,1],y:[0,1]}}
  }}
],{{
  paper_bgcolor:'#fff',plot_bgcolor:'#fff',
  font:{{family:'Inter',color:'#1f2937',size:13}},
  margin:{{l:80,r:80,t:40,b:40}},
  showlegend:false,
  autosize:true,
}},{{responsive:true,displayModeBar:false}});

</script></body></html>"""


def page_complex_dashboard():
    c = compute_complex()
    month_str = f"{st.session_state.month} {st.session_state.year}"

    col_sw, col_ed, col_xl, col_pdf, col_so = st.columns([2, 1, 1, 1, 1])
    with col_sw:
        if st.button("Switch Mode", use_container_width=True, key="sw_cxd"):
            st.session_state.page = "mode_select"
            st.rerun()
    with col_ed:
        if st.button("Edit", use_container_width=True, key="ed_cxd"):
            st.session_state.page = "cx_step"
            st.session_state.cx_step = 0
            st.rerun()
    with col_xl:
        buf = build_complex_excel(c, month_str)
        st.download_button("Download Excel", data=buf,
                           file_name=f"budget_complex_{month_str.replace(' ','_')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="xl_cxd")
    with col_pdf:
        pdf_buf = build_complex_pdf(c, month_str)
        st.download_button("Download PDF", data=pdf_buf,
                           file_name=f"budget_complex_{month_str.replace(' ', '_')}.pdf",
                           mime="application/pdf",
                           use_container_width=True, key="pdf_cxd")
    with col_so:
        if st.button("Start Over", use_container_width=True, key="so_cxd"):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()

    html = build_complex_dashboard(c, month_str)
    components.html(html, height=1150, scrolling=False)


# ══════════════════════════════════════════════════════════════════════════════
# PDF export helpers (reportlab, chart-free)
# ══════════════════════════════════════════════════════════════════════════════
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable)
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

_NAVY_RL  = colors.HexColor("#1e3a5f")
_GREEN_RL = colors.HexColor("#059669")
_RED_RL   = colors.HexColor("#dc2626")
_GREY_RL  = colors.HexColor("#f3f4f6")
_LGREY_RL = colors.HexColor("#eff6ff")


def _pdf_styles():
    s = getSampleStyleSheet()
    title   = ParagraphStyle("Title2",   fontName="Helvetica-Bold", fontSize=14,
                              textColor=_NAVY_RL, spaceAfter=4)
    sub     = ParagraphStyle("Sub",      fontName="Helvetica",      fontSize=9,
                              textColor=colors.HexColor("#6b7280"), spaceAfter=12)
    sec_hdr = ParagraphStyle("SecHdr",   fontName="Helvetica-Bold", fontSize=8,
                              textColor=_NAVY_RL, spaceBefore=6, spaceAfter=2)
    normal  = ParagraphStyle("Normal2",  fontName="Helvetica",      fontSize=9,
                              textColor=colors.HexColor("#111827"))
    bold    = ParagraphStyle("Bold2",    fontName="Helvetica-Bold", fontSize=9,
                              textColor=colors.HexColor("#111827"))
    return dict(title=title, sub=sub, sec_hdr=sec_hdr, normal=normal, bold=bold)


def _fmt_r(v): return f"R {v:,.2f}"
def _var_col(v): return _GREEN_RL if v >= 0 else _RED_RL
def _var_str(v):
    sign = "+" if v >= 0 else "-"
    return f"{sign}R {abs(v):,.2f}"


def _table_style_base():
    return [
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("BACKGROUND",  (0, 0), (-1, 0),  _NAVY_RL),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN",       (0, 0), (0, -1),  "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _GREY_RL]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",(0, 0), (-1, -1), 6),
    ]


def build_simple_pdf(c, month_str):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    st2 = _pdf_styles()
    story = []

    story.append(Paragraph(f"Monthly Budget Review", st2["title"]))
    story.append(Paragraph(f"{month_str} &nbsp;|&nbsp; Simple Mode", st2["sub"]))
    story.append(HRFlowable(width="100%", thickness=1, color=_NAVY_RL, spaceAfter=10))

    col_w = [8*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    data  = [["Category", "Projected", "Actual", "Variance"]]

    def add_sec(label):
        data.append([label, "", "", ""])

    def add_row(label, p, a, bold=False):
        if not bold and p == 0 and a == 0:
            return
        v = a - p
        data.append([label, _fmt_r(p), _fmt_r(a), _var_str(v)])

    add_sec("INCOME")
    add_row("Total Income", c["tip"], c["tia"], bold=True)
    add_sec("OUTFLOWS")
    add_row("Debt Repayments",       c["dp"], c["da"])
    add_row("Fixed Expenses",        c["fp"], c["fa"])
    add_row("Variable Expenses",     c["vp"], c["va"])
    add_row("Savings & Investments", c["sp"], c["sa"])
    add_row("Total Outflows",        c["top_"], c["toa"], bold=True)
    add_row("Net Surplus / Deficit", c["surp_p"], c["surp_a"], bold=True)

    ts = _table_style_base()
    # Section header rows — navy bg
    sec_rows = [i for i, r in enumerate(data) if r[0] in
                ("INCOME","OUTFLOWS")]
    for ri in sec_rows:
        ts += [("BACKGROUND", (0,ri),(-1,ri), _LGREY_RL),
               ("TEXTCOLOR",  (0,ri),(-1,ri), _NAVY_RL),
               ("FONTNAME",   (0,ri),(-1,ri), "Helvetica-Bold"),
               ("SPAN",       (0,ri),(-1,ri))]
    # Colour variance column
    for ri, row in enumerate(data[1:], 1):
        if row[3] and row[3] not in ("", "Variance"):
            v = c["surp_a"] if row[0] == "Net Surplus / Deficit" else 0
            try:
                raw = row[3].replace("R ","").replace(",","").replace("+","").replace("-","")
                num = float(raw)
                sign = 1 if "+" in row[3] else -1
                col = _GREEN_RL if sign > 0 else _RED_RL
                ts.append(("TEXTCOLOR", (3,ri),(3,ri), col))
            except Exception:
                pass
    # Bold total rows
    bold_rows = [i for i, r in enumerate(data) if r[0] in
                 ("Total Income","Total Outflows","Net Surplus / Deficit")]
    for ri in bold_rows:
        ts.append(("FONTNAME", (0,ri),(-1,ri), "Helvetica-Bold"))
        if data[ri][0] == "Net Surplus / Deficit":
            ts.append(("LINEABOVE", (0,ri),(-1,ri), 1, _NAVY_RL))

    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 0.5*cm))

    # Summary section
    story.append(Paragraph("Budget Summary", st2["title"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_NAVY_RL, spaceAfter=6))

    tia = c["tia"]
    surplus_a = c["surp_a"]
    sav_rate  = (c["sa"]/tia*100) if tia > 0 else 0
    req_red   = max(0, -surplus_a)

    sum_data = [
        ["What you should have left",        fmt(surplus_a)],
        ["Savings Rate",                     f"{sav_rate:.1f}% of income" if tia>0 else "—"],
        ["Required Reduction to Break Even", fmt(req_red) if req_red>0 else "None needed"],
    ]
    sum_tbl = Table(sum_data, colWidths=[10*cm, 8.5*cm])
    sum_tbl.setStyle(TableStyle([
        ("FONTSIZE",    (0,0),(-1,-1), 9),
        ("FONTNAME",    (1,0),(1,-1),  "Helvetica-Bold"),
        ("ALIGN",       (1,0),(1,-1),  "RIGHT"),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white, _GREY_RL]),
        ("GRID",        (0,0),(-1,-1), 0.3, colors.HexColor("#e5e7eb")),
        ("TOPPADDING",  (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING", (0,0),(-1,-1), 6),
    ]))
    story.append(sum_tbl)

    doc.build(story)
    buf.seek(0)
    return buf


def build_complex_pdf(c, month_str):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    st2 = _pdf_styles()
    story = []

    story.append(Paragraph("Monthly Budget Review", st2["title"]))
    story.append(Paragraph(f"{month_str} &nbsp;|&nbsp; Complex Mode", st2["sub"]))
    story.append(HRFlowable(width="100%", thickness=1, color=_NAVY_RL, spaceAfter=10))

    debt_lm  = {"home_loan":"Home Loan","vehicle":"Vehicle Finance",
                "credit_card":"Credit Cards","personal_loan":"Personal Loans",
                "student_loan":"Student Loans","other_debt":"Other Debt"}
    fixed_lm = {"rent":"Rent","insurance":"Insurance","medical_aid":"Medical Aid",
                "rates":"Rates / Levies","fibre":"Fibre / Internet","phone":"Phone",
                "school_fees":"School Fees","subscriptions":"Subscriptions",
                "security":"Security","domestic":"Domestic Workers","other_fixed":"Other Fixed"}
    var_lm   = {"groceries":"Groceries","fuel":"Fuel / Transport",
                "electricity":"Electricity / Gas","clothing":"Clothing",
                "entertainment":"Entertainment","restaurants":"Restaurants / Takeaways",
                "alcohol":"Alcohol / Cigarettes","self_care":"Self-Care",
                "travel":"Holidays / Travel","misc":"Miscellaneous"}
    sav_lm   = {"ra":"Retirement Annuities","unit_trusts":"Unit Trusts",
                "tfsa":"Tax-Free Savings","offshore":"Offshore Investments",
                "emergency":"Emergency Fund","other_savings":"Other Savings"}
    inc_lm   = {"salary":"Salary / Primary Income","additional":"Additional Income",
                "rental":"Rental Income","business":"Business Income",
                "other_income":"Other Income"}

    col_w = [9*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    data  = [["Category", "Projected", "Actual", "Variance"]]

    def add_sec(label):
        data.append([label, "", "", ""])

    def add_row(label, p, a, bold=False):
        if not bold and p == 0 and a == 0:
            return
        v = a - p
        data.append([label, _fmt_r(p), _fmt_r(a), _var_str(v)])

    add_sec("INCOME")
    for fk, fl in inc_lm.items():
        add_row(fl, c["inc_p"].get(fk,0), c["inc_a"].get(fk,0))
    add_row("Total Income", c["tip"], c["tia"], bold=True)

    add_sec("DEBT REPAYMENTS")
    for fk, fl in debt_lm.items():
        add_row(fl, c["debt_p"].get(fk,0), c["debt_a"].get(fk,0))
    add_row("Total Debt", c["tdp"], c["tda"], bold=True)

    add_sec("FIXED EXPENSES")
    for fk, fl in fixed_lm.items():
        add_row(fl, c["fixed_p"].get(fk,0), c["fixed_a"].get(fk,0))
    add_row("Total Fixed", c["tfp"], c["tfa"], bold=True)

    add_sec("VARIABLE EXPENSES")
    for fk, fl in var_lm.items():
        add_row(fl, c["var_p"].get(fk,0), c["var_a"].get(fk,0))
    add_row("Total Variable", c["tvp"], c["tva"], bold=True)

    add_sec("SAVINGS & INVESTMENTS")
    for fk, fl in sav_lm.items():
        add_row(fl, c["sav_p"].get(fk,0), c["sav_a"].get(fk,0))
    add_row("Total Savings", c["tsp"], c["tsa"], bold=True)

    add_row("NET SURPLUS / DEFICIT", c["surp_p"], c["surp_a"], bold=True)

    sec_names = {"INCOME","DEBT REPAYMENTS","FIXED EXPENSES",
                 "VARIABLE EXPENSES","SAVINGS & INVESTMENTS"}
    total_names = {"Total Income","Total Debt","Total Fixed","Total Variable",
                   "Total Savings","NET SURPLUS / DEFICIT"}

    ts = _table_style_base()
    for ri, row in enumerate(data):
        if row[0] in sec_names:
            ts += [("BACKGROUND",(0,ri),(-1,ri),_LGREY_RL),
                   ("TEXTCOLOR", (0,ri),(-1,ri),_NAVY_RL),
                   ("FONTNAME",  (0,ri),(-1,ri),"Helvetica-Bold"),
                   ("SPAN",      (0,ri),(-1,ri))]
        if row[0] in total_names:
            ts.append(("FONTNAME",(0,ri),(-1,ri),"Helvetica-Bold"))
        if ri > 0 and len(row) > 3 and row[3]:
            try:
                sign = 1 if row[3].startswith("+") else -1
                col  = _GREEN_RL if sign > 0 else _RED_RL
                ts.append(("TEXTCOLOR",(3,ri),(3,ri),col))
            except Exception:
                pass

    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle(ts))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    # Financial ratios summary
    story.append(Paragraph("Financial Ratios & Advisor Summary", st2["title"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=_NAVY_RL, spaceAfter=6))

    tia      = c["tia"]
    surp     = c["surp_a"]
    sav_rate = (c["tsa"]/tia*100) if tia > 0 else 0
    dti      = (c["tda"]/tia*100) if tia > 0 else 0
    fixed_r  = ((c["tda"]+c["tfa"])/tia*100) if tia > 0 else 0
    ess_p    = (c["ess_a"]/tia*100) if tia > 0 else 0
    non_p    = (c["non_a"]/tia*100) if tia > 0 else 0
    req_red  = max(0, -surp)

    ratio_data = [
        ["Metric",                          "Value",          "Benchmark"],
        ["Income Surplus",                  _fmt_r(surp),     ""],
        ["Savings Rate",                    f"{sav_rate:.1f}%","Target: >20%"],
        ["Debt-to-Income Ratio",            f"{dti:.1f}%",    "Safe: <30%"],
        ["Fixed Obligation Ratio",          f"{fixed_r:.1f}%","<60% recommended"],
        ["Essential Spending",              f"{ess_p:.1f}%",  "Savings + essentials"],
        ["Non-Essential Spending",          f"{non_p:.1f}%",  "Discretionary items"],
        ["Required Reduction to Break Even",_fmt_r(req_red) if req_red>0 else "None needed",""],
    ]
    rtbl = Table(ratio_data, colWidths=[8*cm, 4*cm, 6.5*cm])
    rtbl.setStyle(TableStyle([
        ("FONTNAME",    (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0),(-1,-1), 9),
        ("BACKGROUND",  (0,0),(-1,0),  _NAVY_RL),
        ("TEXTCOLOR",   (0,0),(-1,0),  colors.white),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,_GREY_RL]),
        ("GRID",        (0,0),(-1,-1), 0.3, colors.HexColor("#e5e7eb")),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0),(-1,-1), 6),
        ("RIGHTPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(rtbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# Excel export helpers
# ══════════════════════════════════════════════════════════════════════════════
NAVY_HEX  = "1E3A5F"
GREEN_HEX = "059669"
RED_HEX   = "DC2626"
GREY_HEX  = "F3F4F6"

def _border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_simple_excel(c, month_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Review"

    navy_fill  = PatternFill("solid", fgColor=NAVY_HEX)
    grey_fill  = PatternFill("solid", fgColor="EFF6FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    sec_font   = Font(name="Calibri", bold=True, color=NAVY_HEX, size=10)
    bold_font  = Font(name="Calibri", bold=True, size=10)
    norm_font  = Font(name="Calibri", size=10)
    right_al   = Alignment(horizontal="right")
    left_al    = Alignment(horizontal="left")

    # Title row
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Monthly Budget Review — {month_str} (Simple Mode)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=NAVY_HEX)
    ws["A1"].alignment = left_al
    ws.row_dimensions[1].height = 22
    ws.append([])

    # Header row
    headers = ["Category", "Projected", "Actual", "Variance"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.fill   = navy_fill
        cell.font   = hdr_font
        cell.border = _border()
        cell.alignment = right_al if col > 1 else left_al

    def add_section(label):
        ws.append([label, "", "", ""])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.font   = sec_font
            cell.fill   = grey_fill
            cell.border = _border()

    def add_row(label, p, a, bold=False):
        v = a - p
        ws.append([label, p, a, v])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.fill   = white_fill
            cell.border = _border()
            cell.font   = bold_font if bold else norm_font
            if col > 1:
                cell.alignment = right_al
                if col in (2, 3):
                    cell.number_format = 'R #,##0'
                if col == 4:
                    cell.number_format = 'R #,##0'
                    cell.font = Font(name="Calibri", bold=bold,
                                     color=GREEN_HEX if v >= 0 else RED_HEX, size=10)

    add_section("INCOME")
    add_row("Total Income", c["tip"], c["tia"], bold=True)

    add_section("OUTFLOWS")
    add_row("Debt Repayments",       c["dp"], c["da"])
    add_row("Fixed Expenses",        c["fp"], c["fa"])
    add_row("Variable Expenses",     c["vp"], c["va"])
    add_row("Savings & Investments", c["sp"], c["sa"])
    add_row("Total Outflows",        c["top_"], c["toa"], bold=True)
    add_row("Net Surplus / Deficit", c["surp_p"], c["surp_a"], bold=True)

    _set_col_widths(ws, [28, 16, 16, 16])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_complex_excel(c, month_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Review"

    navy_fill  = PatternFill("solid", fgColor=NAVY_HEX)
    grey_fill  = PatternFill("solid", fgColor="EFF6FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_font   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    sec_font   = Font(name="Calibri", bold=True, color=NAVY_HEX, size=10)
    bold_font  = Font(name="Calibri", bold=True, size=10)
    norm_font  = Font(name="Calibri", size=10)
    right_al   = Alignment(horizontal="right")
    left_al    = Alignment(horizontal="left")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Monthly Budget Review — {month_str} (Complex Mode)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=NAVY_HEX)
    ws["A1"].alignment = left_al
    ws.row_dimensions[1].height = 22
    ws.append([])

    headers = ["Category", "Projected", "Actual", "Variance"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.fill   = navy_fill
        cell.font   = hdr_font
        cell.border = _border()
        cell.alignment = right_al if col > 1 else left_al

    def add_section(label):
        ws.append([label, "", "", ""])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.font   = sec_font
            cell.fill   = grey_fill
            cell.border = _border()

    def add_row(label, p, a, bold=False):
        v = a - p
        ws.append([label, p, a, v])
        r = ws.max_row
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.fill   = white_fill
            cell.border = _border()
            cell.font   = bold_font if bold else norm_font
            if col > 1:
                cell.alignment = right_al
                if col in (2, 3):
                    cell.number_format = 'R #,##0'
                if col == 4:
                    cell.number_format = 'R #,##0'
                    cell.font = Font(name="Calibri", bold=bold,
                                     color=GREEN_HEX if v >= 0 else RED_HEX, size=10)

    debt_labels  = {"home_loan":"Home Loan","vehicle":"Vehicle Finance",
                    "credit_card":"Credit Cards","personal_loan":"Personal Loans",
                    "student_loan":"Student Loans","other_debt":"Other Debt"}
    fixed_labels = {"rent":"Rent","insurance":"Insurance","medical_aid":"Medical Aid",
                    "rates":"Rates / Levies","fibre":"Fibre / Internet","phone":"Phone",
                    "school_fees":"School Fees","subscriptions":"Subscriptions",
                    "security":"Security","domestic":"Domestic Workers","other_fixed":"Other Fixed"}
    var_labels   = {"groceries":"Groceries","fuel":"Fuel / Transport",
                    "electricity":"Electricity / Gas","clothing":"Clothing",
                    "entertainment":"Entertainment","restaurants":"Restaurants / Takeaways",
                    "alcohol":"Alcohol / Cigarettes","self_care":"Self-Care",
                    "travel":"Holidays / Travel","misc":"Miscellaneous"}
    sav_labels   = {"ra":"Retirement Annuities","unit_trusts":"Unit Trusts",
                    "tfsa":"Tax-Free Savings","offshore":"Offshore Investments",
                    "emergency":"Emergency Fund","other_savings":"Other Savings"}
    inc_labels   = {"salary":"Salary / Primary Income","additional":"Additional Income",
                    "rental":"Rental Income","business":"Business Income","other_income":"Other Income"}

    add_section("INCOME")
    for fk, fl in inc_labels.items():
        add_row(fl, c["inc_p"].get(fk,0), c["inc_a"].get(fk,0))
    add_row("Total Income", c["tip"], c["tia"], bold=True)

    add_section("DEBT REPAYMENTS")
    for fk, fl in debt_labels.items():
        add_row(fl, c["debt_p"].get(fk,0), c["debt_a"].get(fk,0))
    add_row("Total Debt", c["tdp"], c["tda"], bold=True)

    add_section("FIXED EXPENSES")
    for fk, fl in fixed_labels.items():
        add_row(fl, c["fixed_p"].get(fk,0), c["fixed_a"].get(fk,0))
    add_row("Total Fixed", c["tfp"], c["tfa"], bold=True)

    add_section("VARIABLE EXPENSES")
    for fk, fl in var_labels.items():
        add_row(fl, c["var_p"].get(fk,0), c["var_a"].get(fk,0))
    add_row("Total Variable", c["tvp"], c["tva"], bold=True)

    add_section("SAVINGS & INVESTMENTS")
    for fk, fl in sav_labels.items():
        add_row(fl, c["sav_p"].get(fk,0), c["sav_a"].get(fk,0))
    add_row("Total Savings", c["tsp"], c["tsa"], bold=True)

    add_row("Net Surplus / Deficit", c["surp_p"], c["surp_a"], bold=True)

    _set_col_widths(ws, [30, 16, 16, 16])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# Router
# ══════════════════════════════════════════════════════════════════════════════
page = st.session_state.page
if   page == "mode_select":   page_mode_select()
elif page == "month":         page_month()
elif page == "step":          page_simple_step()
elif page == "dashboard":     page_simple_dashboard()
elif page == "cx_step":       page_complex_step()
elif page == "cx_dashboard":  page_complex_dashboard()
